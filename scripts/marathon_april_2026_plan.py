#!/usr/bin/env python3
"""Генерация цветного Excel: тренировки + меню 11–26.04.2026."""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = _ROOT / "half_marathon_plan_aprel_2026.xlsx"

# Цвета
HEADER_FILL = "2F5496"
HEADER_FONT = "FFFFFF"
TITLE_FILL = "1F3864"
ACCENT_ROW = "E2EFDA"  # лёгкий зелёный — обычные дни
WEEKEND_FILL = "FFF2CC"  # выходные
RACE_FILL = "C5E0B4"  # день старта
AFTER_RACE_FILL = "FCE4D6"  # восстановление
ALT_ROW = "F2F2F2"

ROWS = [
    {
        "d": date(2026, 4, 11),
        "train": "Лёгкая зарядка и растяжка 15–20 мин; пресс 15 мин (планка, подъёмы ног)",
        "z": "Овсянка 60 г (сух.) на воде/молоке 200 мл, банан 1 небольшой, творог 5% 150 г, чай без сахара.",
        "o": "Гречка 180 г, куриная грудка 200 г, салат (огурец, помидор, зелень) 250 г + 1 ч. л. масла, кефир 1% 200 мл.",
        "u": "Рыба 200 г, тушёная капуста/брокколи 300 г, хлеб цельнозерновой 30 г.",
        "p": "Яблоко 1 шт.",
        "kind": "weekend",
    },
    {
        "d": date(2026, 4, 12),
        "train": "Лёгкий бег 6–8 км спокойно; после — 10 мин пресса",
        "z": "Яичница 3 яйца + 2 белка, хлеб 30 г, помидор/огурец 150 г.",
        "o": "Индейка/говядина 180 г, булгур/рис 150 г, салат 200 г, масло 1 ч. л.",
        "u": "Творог 5% 200 г, сметана 10% 1 ст. л., ягоды 80 г или мандарин.",
        "p": "Греческий йогурт 150 г.",
        "kind": "weekend",
    },
    {
        "d": date(2026, 4, 13),
        "train": "Отдых или прогулка 30–40 мин; лёгкая растяжка",
        "z": "Запеканка: творог 200 г + 1 яйцо + 1 ст. л. манки (или просто творог 200 г + ½ банана).",
        "o": "Рыба 200 г, картофель запечённый 150 г, салат 250 г, масло 1 ч. л.",
        "u": "Салат с куриной грудкой 150 г, хлебцы 2 шт., йогурт 100 г.",
        "p": "Творог 100 г или батончик ~200 ккал.",
        "kind": "normal",
    },
    {
        "d": date(2026, 4, 14),
        "train": "Бег 8–10 км лёгко; пресс 15 мин после",
        "z": "Овсянка 70 г, молоко 200 мл, мёд 1 ч. л., яйца варёные 2 шт.",
        "o": "Паста 70 г (сух.) + курица 180 г + томатный соус, салат 150 г.",
        "u": "Рыба 180 г, гречка 120 г, овощи на пару 250 г.",
        "p": "Банан 1 шт. (до/после бега по самочувствию).",
        "kind": "normal",
    },
    {
        "d": date(2026, 4, 15),
        "train": "Сила: подтягивания 4–6×, брусья, приседы/выпады; пресс 20 мин",
        "z": "Сырники из 200 г творога, сметана 1 ст. л., ягоды 50 г.",
        "o": "Суп с фрикадельками (~150 г мяса), картофель 1 шт., морковь; хлеб 25 г.",
        "u": "Индейка 200 г, кабачки тушёные 300 г, гречка 100 г.",
        "p": "Огурец + помидор или яблоко.",
        "kind": "normal",
    },
    {
        "d": date(2026, 4, 16),
        "train": "Бег 10–12 км + 3–4×3 мин чуть быстрее комфортного темпа",
        "z": "Гречка 150 г готовой + творог 150 г + 1 ст. л. сметаны.",
        "o": "Тунец в соку 1 банка, яйца 2 шт., фасоль стручковая 100 г, листья, масло 1 ч. л., хлебцы 2 шт.",
        "u": "Куриная грудка 200 г, рис 130 г, морковь/брокколи 200 г.",
        "p": "Кефир 250 мл.",
        "kind": "normal",
    },
    {
        "d": date(2026, 4, 17),
        "train": "Отдых / ходьба 40 мин; без силовой на отказ",
        "z": "Омлет 3 яйца + молоко 50 мл, сыр 30 г, помидор.",
        "o": "Котлета куриная на пару 150 г, пюре 150 г, салат 200 г.",
        "u": "Рыба 180 г, салат с авокадо 40 г (или оливки 5 шт.), овощи 200 г.",
        "p": "Творог 120 г.",
        "kind": "normal",
    },
    {
        "d": date(2026, 4, 18),
        "train": "Длинный бег 14–16 км спокойно; растяжка 10 мин",
        "z": "Овсянка 75 г, банан 1, молоко 200 мл, орехи 15 г.",
        "o": "После бега: суп с лапшой (лапша 50 г сух., курица 150 г) ИЛИ гречка 200 г + грудка 180 г + овощи.",
        "u": "Творог 5% 250 г, кефир 200 мл, ягоды или 1 ч. л. мёда.",
        "p": "Финики 2–3 или банан при сильном голоде.",
        "kind": "weekend",
    },
    {
        "d": date(2026, 4, 19),
        "train": "Восстановление: бег 4–6 км очень лёго или вело/ходьба 40 мин",
        "z": "Творог 200 г, ягоды, 1 ст. л. йогурта.",
        "o": "Гречка 180 г, рыба 200 г, салат 250 г, масло 1 ч. л.",
        "u": "Рагу с курицей 250 г (курица + овощи), хлебец 1 шт.",
        "p": "Яблоко.",
        "kind": "weekend",
    },
    {
        "d": date(2026, 4, 20),
        "train": "Лёгкий бег 6–8 км",
        "z": "Хлеб 50 г, ветчина низкой жирности 80 г (или курица 100 г), огурец, салат; чай.",
        "o": "Бефстроганов говядина 180 г, гречка 150 г, квашеная капуста 100 г.",
        "u": "Рыба запечённая 200 г, цветная капуста 250 г.",
        "p": "Йогурт греческий 120 г.",
        "kind": "normal",
    },
]

ROWS.extend(
    [
        {
            "d": date(2026, 4, 21),
            "train": "Сила + пресс (короче: 2 подхода ключевых упражнений)",
            "z": "Овсянка 65 г, молоко 200 мл, ягоды.",
            "o": "Салат: курица 200 г, фасоль красная 4 ст. л., кукуруза 2 ст. л., йогурт+горчица.",
            "u": "Индейка 200 г, баклажан/кабачки 250 г, рис 100 г.",
            "p": "Творог 100 г.",
            "kind": "normal",
        },
        {
            "d": date(2026, 4, 22),
            "train": "Бег 8–10 км + 4–6 коротких ускорений по 20–30 с",
            "z": "Яйца 3 варёных, авокадо 50 г или огурец 200 г, хлебец 1 шт.",
            "o": "Рыбные котлеты 200 г, картофель 150 г, салат 200 г.",
            "u": "Суп-пюре куриный 400 мл, крекеры 20 г.",
            "p": "Кефир 200 мл.",
            "kind": "normal",
        },
        {
            "d": date(2026, 4, 23),
            "train": "Отдых; лёгкая прогулка; растяжка",
            "z": "Сырники 200 г творога + яйцо, сметана 1 ст. л.",
            "o": "Паста 60 г сухая, фарш 150 г в томатах, салат.",
            "u": "Рыба 180 г, гречка 120 г, овощи 200 г.",
            "p": "½ банана.",
            "kind": "normal",
        },
        {
            "d": date(2026, 4, 24),
            "train": "Бег 6–8 км лёгко + растяжка; рано лечь спать",
            "z": "Овсянка 80 г, мёд 1 ст. л., банан 1, молоко 200 мл.",
            "o": "Рис 200 г готовый (или паста 70 г сух.), курица 150 г, овощи 200 г без жира.",
            "u": "Рыба/курица 150 г, картофель 200 г или паста 50 г сух., огурец/помидор 200 г.",
            "p": "Сухофрукты 30–40 г или банан; вода весь день.",
            "kind": "normal",
        },
        {
            "d": date(2026, 4, 25),
            "train": "ПОЛУМАРАФОН 21,1 км — удачи!",
            "z": "За 3 ч до старта: овсянка 50–60 г на воде, ½–1 банан, мёд 1 ч. л., вода 200–300 мл.",
            "o": "Во время — по вашему плану гелей/воды. После финиша: вода + банан/йогурт/бутерброд.",
            "u": "Ужин после гонки: суп + второе ИЛИ 2 куска тонкой пиццы + салат; белок (мясо/рыба/творог).",
            "p": "По желанию после ужина.",
            "kind": "race",
        },
        {
            "d": date(2026, 4, 26),
            "train": "Отдых: прогулка 30–40 мин, растяжка; без интенсива",
            "z": "Оладьи из овсянки (овсянка 60 г + яйцо + банан) или блины с творогом; ягоды.",
            "o": "Борщ без сала + мясо 100 г + картофель 1 шт. или гречка 150 г, сметана 1 ст. л.",
            "u": "Рыба 200 г, салат 250 г, хлеб 40 г.",
            "p": "Фрукт, орехи 20 г.",
            "kind": "after_race",
        },
    ]
)

WD_RU = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "План апрель 2026"

    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Заголовок листа
    ws.merge_cells("A1:H1")
    c1 = ws["A1"]
    c1.value = (
        "Полумарафон 25.04.2026  •  Цель: −3…5 кг  •  Белок ~150–165 г/день  •  Вода 2–2,5 л"
    )
    c1.font = title_font
    c1.fill = PatternFill("solid", fgColor=TITLE_FILL)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Дата",
        "День",
        "Тренировка",
        "Завтрак",
        "Обед",
        "Ужин",
        "Перекус",
        "Тип дня",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[2].height = 22

    widths = [11, 6, 38, 36, 36, 36, 32, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, row in enumerate(ROWS):
        rnum = 3 + idx
        wd = WD_RU[row["d"].weekday()].capitalize()
        kind = row["kind"]
        if kind == "weekend":
            fill = WEEKEND_FILL
            day_label = "выходной"
        elif kind == "race":
            fill = RACE_FILL
            day_label = "СТАРТ"
        elif kind == "after_race":
            fill = AFTER_RACE_FILL
            day_label = "после гонки"
        else:
            fill = ACCENT_ROW if idx % 2 == 0 else ALT_ROW
            day_label = "будни"

        vals = [
            row["d"].strftime("%d.%m.%Y"),
            wd,
            row["train"],
            row["z"],
            row["o"],
            row["u"],
            row["p"],
            day_label,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=rnum, column=col, value=val)
            cell.font = cell_font
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = wrap
            cell.border = border
        ws.row_dimensions[rnum].height = 96

    last_data_row = 2 + len(ROWS)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{last_data_row}"

    leg_row = last_data_row + 1
    ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=8)
    leg = ws.cell(row=leg_row, column=1)
    leg.value = (
        "Легенда цветов строк:  жёлтый — суббота/воскресенье  |  "
        "зелёный/серый чередование — будни  |  "
        "насыщенный зелёный — день полумарафона  |  "
        "персиковый — день после старта."
    )
    leg.font = Font(name="Calibri", size=9, italic=True, color="404040")
    leg.fill = PatternFill("solid", fgColor="DEEAF6")
    leg.alignment = Alignment(wrap_text=True, vertical="center")
    leg.border = border
    ws.row_dimensions[leg_row].height = 36

    # Второй лист — краткая шпаргалка по тренировкам
    ws2 = wb.create_sheet("Тренировки кратко")
    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value = "Недельный фокус"
    t.font = title_font
    t.fill = PatternFill("solid", fgColor=TITLE_FILL)
    t.alignment = Alignment(horizontal="center")
    tips = [
        ("Бег", "80% объёма — спокойный разговорный темп; не гонитесь за скоростью на длинных."),
        ("Сила", "Подтягивания и брусья — до 2 повторов в запасе; пресс ежедневно по 15–20 мин в лёгие дни."),
        ("Сон", "7–8 ч; перед стартом ложитесь раньше 2 ночи подряд."),
        ("Вес", "Взвешивание 1× в неделю утром; при слабости на бегу — +углеводы в завтрак."),
    ]
    tip_fill_a = PatternFill("solid", fgColor="B4C6E7")
    tip_fill_b = PatternFill("solid", fgColor="D9E1F2")
    for i, (a, b) in enumerate(tips, 2):
        ca = ws2.cell(row=i, column=1, value=a)
        ca.font = Font(name="Calibri", size=11, bold=True, color="1F3864")
        ca.fill = tip_fill_a
        ca.alignment = Alignment(vertical="top", horizontal="center")
        ca.border = border
        cb = ws2.cell(row=i, column=2, value=b)
        cb.font = cell_font
        cb.fill = tip_fill_b
        cb.alignment = wrap
        cb.border = border
        ws2.row_dimensions[i].height = 48
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 72
    ws2.row_dimensions[1].height = 26

    wb.save(OUT_PATH)
    print("Saved", OUT_PATH)


if __name__ == "__main__":
    build()
